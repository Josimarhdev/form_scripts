from openpyxl import load_workbook, Workbook  # Para trabalhar com arquivos excel
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # Para aplicas estilos nas cédulas
from datetime import datetime  # Para manipular datas
import pandas as pd
from pathlib import Path  # Para manipulação de caminhos de arquivos
from utils import (  #Estilos e funções auxiliares
    cabeçalho_fill, cabeçalho_font, enviado_fill, enviado_font,
    semtecnico_fill, atrasado_fill, validado_nao_fill, validado_sim_fill,
    cores_regionais, bordas, alinhamento,
    normalizar_texto, normalizar_uvr, aplicar_estilo_status
)


# Define o caminho do script atual
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_form2 = pasta_scripts.parent / "form2"

# Caminho do arquivo do banco e arquivos auxiliares (originais do drive)
csv_file_input = pasta_form2 / "planilhas_consumo/form2.csv"  
planilhas_auxiliares = {
    "belem": pasta_form2 / "planilhas_consumo/belem.xlsx",
    "expansao": pasta_form2 / "planilhas_consumo/expansao.xlsx",
    "grs": pasta_form2 / "planilhas_consumo/GRS.xlsx"
}


# Carrega a planilha principal
df_input = pd.read_csv(csv_file_input, dtype=str)

# Dicionário para armazenar o status de envio por município
dados_atualizados = {}

for _, row in df_input.iterrows():
    municipio = row['municipio']
    uvr_nro = row['uvr_nro']
    data_envio = row['data_envio']

    # Normaliza o nome do município
    if isinstance(municipio, str):
        municipio_uvr_normalizado = f"{normalizar_texto(municipio)}_{uvr_nro}"
    else:
        continue  

    # Formata a data de envio
    if isinstance(data_envio, datetime):
        data_envio_formatada = data_envio.strftime("%d/%m/%Y")
    else:
        try:
            data_envio_formatada = datetime.strptime(data_envio, "%Y-%m-%d %H:%M:%S.%f").strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_envio_formatada = ""

    # Verifica se já há entrada para o município/UVR e marca como duplicado se for o caso
    if municipio_uvr_normalizado in dados_atualizados:
        dados_atualizados[municipio_uvr_normalizado]["datas"].append(data_envio_formatada)
        dados_atualizados[municipio_uvr_normalizado]["status"] = "Duplicado"
        print('entrou if')
    else:
        dados_atualizados[municipio_uvr_normalizado] = {
            "datas": [data_envio_formatada],
            "status": "Enviado"
        }
        print('else')


# Processa cada uma das planilhas auxiliares (Belém, GRS e Expansão)
for nome, caminho in planilhas_auxiliares.items():
    wb_aux = load_workbook(caminho)

    # Verifica se a aba correta existe
    if "Form 2 - UVR" not in wb_aux.sheetnames:
        print(f"A aba 'Form 2 - UVR' não foi encontrada em {nome}. Nenhuma modificação será feita.")
        continue

    ws_aux = wb_aux["Form 2 - UVR"]

    # Cria uma nova planilha para salvar os dados atualizados
    novo_wb = Workbook()
    novo_ws = novo_wb.active
    novo_ws.title = nome

    # Copia e estiliza os cabeçalhos
    headers = [cell.value for cell in ws_aux[1]]
    for col_num, header in enumerate(headers, start=1):
        cell = novo_ws.cell(row=1, column=col_num, value=header)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # Processa cada linha da planilha auxiliar
    for row_idx, row in enumerate(ws_aux.iter_rows(min_row=2, values_only=True), start=2):
        municipio_original = row[1]
        uvr_nro_original = row[2]
        row_data = list(row)

        # Normaliza o nome do município e UVR, fazendo a concatenação para comparação posterior
        if isinstance(municipio_original, str):
            municipio_uvr_normalizado = f"{normalizar_texto(municipio_original)}_{normalizar_uvr(uvr_nro_original)}"
        else:
            municipio_uvr_normalizado = ""

        # Se foi enviado, atualiza status e data com os dados atualizados do drive
        if municipio_uvr_normalizado in dados_atualizados:
            print('deu certo')
            novas_datas = ", ".join(dados_atualizados[municipio_uvr_normalizado]["datas"])
            novo_status = dados_atualizados[municipio_uvr_normalizado]["status"]
            row_data[5] = novas_datas  # Coluna de data de envio
            row_data[4] = novo_status  # Coluna de status
        else:
            # Define como "Atrasado" caso não tenha status
            if row_data[4] == "Sem Técnico":
                pass
            elif row_data[4] is None:
                row_data[4] = "Atrasado"

        ###ESTILIZAÇÃO

        # Coloração de validação (Sim/Não)
        if row_data[6] == "Não":
            novo_ws.cell(row=row_idx, column=7).fill = validado_nao_fill
        elif row_data[6] == "Sim":
            novo_ws.cell(row=row_idx, column=7).fill = validado_sim_fill

        # Coloração da cédula de regional
        regional = row_data[0]
        if regional in cores_regionais:
            cor_hex = cores_regionais[regional]
            novo_ws.cell(row=row_idx, column=1).fill = PatternFill(
                start_color=cor_hex, end_color=cor_hex, fill_type="solid"
            )

        # Escreve os dados na nova planilha com estilos
        for col_idx, value in enumerate(row_data, start=1):
            cell = novo_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = bordas
            cell.alignment = alinhamento
            cell.font = Font(name='Arial', size=11)

    # Aplica estilização com base no status 
    for row_idx in range(2, novo_ws.max_row + 1):
        status_cell = novo_ws.cell(row=row_idx, column=5)
        status = status_cell.value
        aplicar_estilo_status(status_cell, status_cell.value)

    # Ajusta a largura das colunas com base no conteúdo
    for col in novo_ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        novo_ws.column_dimensions[col_letter].width = max_length + 5

    # Salva a nova planilha atualizada
    novo_caminho = pasta_form2 / f"{nome}_atualizado_form2.xlsx"
    novo_wb.save(novo_caminho)
    print(f"{novo_caminho} gerada com sucesso")
