# Importações necessárias
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
import pandas as pd
from datetime import timedelta
from utils import (
    cabeçalho_fill, cabeçalho_font, enviado_fill, analise_fill, enviado_font,
    semtecnico_fill, atrasado_fill, validado_nao_fill, validado_sim_fill, duplicado_fill, outras_fill, atrasado2_fill,
    cores_regionais, bordas, alinhamento,
    normalizar_texto, normalizar_uvr, aplicar_estilo_status
)

# Define os caminhos dos arquivos envolvidos
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"


# Caminho do arquivo do banco e arquivos auxiliares (originais do drive)
csv_file_input = pasta_inputs/"form4.csv"
media_file_input = pasta_inputs/"form4-médias.csv"
planilhas_auxiliares = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 4.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 4.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 4.xlsx"
}

# Carrega a planilha principal
df_input = pd.read_csv(csv_file_input, dtype=str)
df_medias = pd.read_csv(media_file_input, dtype=str)

# Dicionários para armazenar os dados extraídos
dados_atualizados = {}
div_por_municipio = {}
regionais_por_municipio = {}
dados_medias = {}

# Processa e armazena as médias
col_municipio_media = df_medias.columns[0]
col_uvr_media = df_medias.columns[2]
col_media_s1 = df_medias.columns[4]
col_media_s2 = df_medias.columns[12]

for _, row in df_medias.iterrows():
    municipio = row[col_municipio_media]
    uvr_nro = row[col_uvr_media]
    if pd.notna(municipio) and pd.notna(uvr_nro):
        chave = (normalizar_texto(str(municipio)), normalizar_uvr(str(uvr_nro)))
        dados_medias[chave] = {
            "media_s1": pd.to_numeric(row[col_media_s1], errors='coerce'),
            "media_s2": pd.to_numeric(row[col_media_s2], errors='coerce')
        }


# Converte a data de referência para o formato "MM.AA"
def converter_data_para_mes_ano(data_referencia):
    if isinstance(data_referencia, datetime):
        return data_referencia.strftime("%m.%y")
    else:
        try:
            return datetime.strptime(data_referencia, "%Y-%m-%d").strftime("%m.%y")
        except (ValueError, TypeError):
            return ""

# Remove caracteres inválidos do nome da aba
def limpar_nome_aba(nome):
    return nome.replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-").replace("?", "-").replace("[", "").replace("]", "")


for _, row in df_input.iterrows():
    municipio = row['gm_nome']
    uvr_nro = row['guvr_numero']
    data_envio = row['data_de_envio']
    tc_uvr = row['nome_tc_uvr']  
    data_referencia = row['data_de_referencia']
    valor_envio = pd.to_numeric(row[df_input.columns[6]], errors='coerce') 


    if isinstance(municipio, str):
        municipio_uvr_normalizado = f"{normalizar_texto(municipio)}_{uvr_nro}"
    else:
        continue

    mes_ano = converter_data_para_mes_ano(data_referencia)

    # Tenta formatar a data de envio
    if isinstance(data_envio, datetime):
        data_envio_formatada = data_envio.strftime("%d/%m/%Y")
    else:
        try:
            data_envio_formatada = datetime.strptime(data_envio, "%Y-%m-%d").strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_envio_formatada = ""

    # Agrupa dados por município + UVR + mês/ano
    chave = (municipio_uvr_normalizado, mes_ano)
    if chave in dados_atualizados:
        dados_atualizados[chave]["datas_envio"].append(data_envio_formatada)
        dados_atualizados[chave]["status"] = "Duplicado"
    else:
        dados_atualizados[chave] = {
            "datas_envio": [data_envio_formatada],
            "status": "Enviado",
            "municipio_original": municipio,
            "uvr_nro": uvr_nro,
            "mes_ano": mes_ano,
            "tc_uvr" : tc_uvr,
            "valor_envio": valor_envio,
            "data_referencia_dt": pd.to_datetime(data_referencia, errors='coerce')
        }

# Cria um novo workbook para cada planilha auxiliar
wb_final = {nome: Workbook() for nome in planilhas_auxiliares}
for nome in wb_final:
    wb_final[nome].remove(wb_final[nome].active)

# Processa cada planilha auxiliar
for nome, caminho in planilhas_auxiliares.items():
    wb_aux = load_workbook(caminho)

    abas_para_copiar = ["Resumo", "Monitoramento", "Regionais"]

    for nome_aba in abas_para_copiar:
        if nome_aba in wb_aux.sheetnames: #verifica se existe, faz isso em todos (grs,expansao,belem)
            print(f"Copiando aba '{nome_aba}' para o arquivo de '{nome}'...")
            ws_origem = wb_aux[nome_aba]
            ws_destino = wb_final[nome].create_sheet(nome_aba)

            # Copia os dados e estilos célula por célula
            for row in ws_origem.iter_rows():
                for cell in row:
                    new_cell = ws_destino.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                        new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                        new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                        new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                        new_cell.number_format = cell.number_format

            # Copia as dimensões das colunas e linhas
            for col_letter, dim in ws_origem.column_dimensions.items():
                ws_destino.column_dimensions[col_letter].width = dim.width
            for row_index, dim in ws_origem.row_dimensions.items():
                ws_destino.row_dimensions[row_index].height = dim.height

            # Copia as células mescladas
            for merged_cell_range in ws_origem.merged_cells.ranges:
                ws_destino.merge_cells(str(merged_cell_range))
            
            for dv in ws_origem.data_validations.dataValidation:
                ws_destino.add_data_validation(dv)

            for range_string in ws_origem.conditional_formatting:
                rules_list = ws_origem.conditional_formatting[range_string]
                    
                for rule in rules_list:
                    ws_destino.conditional_formatting.add(range_string, rule)

    for aba in wb_aux.sheetnames:
        # Só processa abas no formato MM.AA
        if aba.count('.') == 1 and all(x.isdigit() for x in aba.split('.')):
            mes_ano_aux = aba
            ws_aux = wb_aux[aba]

            mes_ano_limpo = limpar_nome_aba(mes_ano_aux)
            if mes_ano_limpo not in wb_final[nome].sheetnames:
                wb_final[nome].create_sheet(title=mes_ano_limpo)

            ws_final = wb_final[nome][mes_ano_limpo]

            dv_sim_nao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
            ws_final.add_data_validation(dv_sim_nao)

            dv_sim_nao_ti = DataValidation(type="list", formula1='"Sim,Não,Em Análise"', allow_blank=True)
            ws_final.add_data_validation(dv_sim_nao_ti)

            dv_status = DataValidation(type="list", formula1='"Enviado, Atrasado, Atrasado >= 2, Outras Ocorrências, Sem Técnico, Duplicado"', allow_blank=True) #dropdown com sim e nao
            ws_final.add_data_validation(dv_status)           

            # Copia cabeçalhos com formatação
            headers = [cell.value for cell in ws_aux[1]]
            for col_num, header in enumerate(headers, start=1):
                cell = ws_final.cell(row=1, column=col_num, value=header)
                cell.fill = cabeçalho_fill
                cell.font = cabeçalho_font
                cell.border = bordas
                cell.alignment = alinhamento

            ws_final.auto_filter.ref = f"A1:G1"

            # Calcula mês/ano atual
            hoje = datetime.today()
            mes_atual = hoje.month
            ano_atual = hoje.year

            # Função auxiliar para calcular a diferença de meses
            def diferenca_em_meses(ano_alvo, mes_alvo, ano_base, mes_base):
                return (ano_base - ano_alvo) * 12 + (mes_base - mes_alvo)

            # Processa linhas de dados
            for row_idx, row in enumerate(ws_aux.iter_rows(min_row=2, values_only=True), start=2):
                regional = row[0]
                municipio_original = row[1]
                uvr_nro_original = row[2]
                row_data = list(row)


                if aba != '01.25':
                    formula = (
                        f'=IFERROR(IF(INDEX(\'01.25\'!D2:D500, '
                        f'MATCH(B{row_idx}&C{row_idx}, INDEX(\'01.25\'!B2:B500&\'01.25\'!C2:C500, 0), 0))="", "", '
                        f'INDEX(\'01.25\'!D2:D500, '
                        f'MATCH(B{row_idx}&C{row_idx}, INDEX(\'01.25\'!B2:B500&\'01.25\'!C2:C500, 0), 0))), "")'
                    )

                    
                    row_data[3] = formula
               
                if not isinstance(municipio_original, str) or not municipio_original.strip():
                    continue

                municipio_uvr_normalizado = f"{normalizar_texto(municipio_original)}_{normalizar_uvr(uvr_nro_original)}"
                chave_busca = (municipio_uvr_normalizado, mes_ano_aux)
                div_por_municipio[municipio_uvr_normalizado] = nome
                regionais_por_municipio[municipio_uvr_normalizado] = regional

                situacao_atual = row_data[4]
                tem_envio_existente = bool(row_data[5]) and isinstance(row_data[5], str) and row_data[5].strip()

                # Atualiza dados conforme a planilha principal
                if chave_busca in dados_atualizados:
                    row_data[5] = ", ".join(dados_atualizados[chave_busca]["datas_envio"])
                    row_data[4] = dados_atualizados[chave_busca]["status"]
                elif not tem_envio_existente:
                    if situacao_atual not in ("UVR Sem Técnico", "Outras Ocorrências"):
                        try:
                            aba_mes, aba_ano = map(int, mes_ano_aux.split("."))
                            aba_ano += 2000
                        except:
                            aba_mes, aba_ano = None, None

                        if aba_ano and aba_mes:
                            diff = diferenca_em_meses(aba_ano, aba_mes, ano_atual, mes_atual)
                            if diff == 1:
                                row_data[4] = "Atrasado"
                            elif diff >= 2:
                                row_data[4] = "Atrasado >= 2"

                # Estiliza as linhas
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_final.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = bordas
                    cell.alignment = alinhamento
                    cell.font = Font(name='Arial', size=11)
                    
                    if col_idx == 7:
                        dv_sim_nao.add(cell.coordinate)

                    if col_idx == 10:
                        dv_sim_nao_ti.add(cell.coordinate)

                    if col_idx == 5:
                        dv_status.add(cell.coordinate) 
                    

                # Aplica cor para célula de validação
                if row_data[6] == "Não":
                    ws_final.cell(row=row_idx, column=7).fill = validado_nao_fill
                elif row_data[6] == "Sim":
                    ws_final.cell(row=row_idx, column=7).fill = validado_sim_fill

                # Aplica cor da regional
                if regional in cores_regionais:
                    cor_hex = cores_regionais[regional]
                    ws_final.cell(row=row_idx, column=1).fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

            # Aplica estilo com base no status
            for row_idx in range(2, ws_final.max_row + 1):
                status_cell = ws_final.cell(row=row_idx, column=5)
                status = status_cell.value
                aplicar_estilo_status(status_cell, status)

            # Ajusta largura das colunas
            for col in ws_final.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = col[0].column_letter
                wb_final[nome][mes_ano_limpo].column_dimensions[col_letter].width = max_length + 5

            coluna_validado_regional = f"G2:G{ws_final.max_row}" # Coluna G é a 7ª coluna
            coluna_validado_ti = f"J2:J{ws_final.max_row}" # Coluna G é a 10ª coluna
            coluna_status = f"E2:E{ws_final.max_row}" # Coluna E é a 5ª coluna

            rule_sim = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
            ws_final.conditional_formatting.add(coluna_validado_regional, rule_sim) #Se for selecionado Sim, pinta de verde
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_sim) #Se for selecionado Sim, pinta de verde

            rule_nao = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)
            ws_final.conditional_formatting.add(coluna_validado_regional, rule_nao) #Se for selecionado Não, pinta de vermelho
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_nao) #Se for selecionado Sim, pinta de verde

            rule_analise = CellIsRule(operator='equal', formula=['"Em Análise"'], stopIfTrue=True, fill=analise_fill)  
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_analise)    

            status_rules = {
            "Enviado": {"fill": enviado_fill, "font": enviado_font},
            "Atrasado": {"fill": atrasado_fill, "font": enviado_font},
            "Atrasado >= 2": {"fill": atrasado2_fill, "font": enviado_font},
            "Outras Ocorrências": {"fill": outras_fill, "font": enviado_font},
            "Sem Técnico": {"fill": semtecnico_fill, "font": enviado_font},
            "Duplicado": {"fill": duplicado_fill, "font": enviado_font}
        }

            for status_text, styles in status_rules.items():
                rule = CellIsRule(operator='equal',
                                formula=[f'"{status_text}"'],
                                stopIfTrue=True,
                                fill=styles["fill"],
                                font=styles["font"])
                ws_final.conditional_formatting.add(coluna_status, rule) 

            ws_final.freeze_panes = 'D1' #Congela as colunas A,B,C  
            ws_final.column_dimensions['D'].width = 45



 # processa a aba de irregulares (grs,expansao e belem)
for nome, wb in wb_final.items():

    print(nome)
    chaves_existentes = set()
    
    caminho_aux = planilhas_auxiliares[nome]
    wb_aux = load_workbook(caminho_aux)

    # cria a aba de irregulares no arquivo final (ela sempre é recriada, porém coletando as informações já existentes na planilha de entrada)
    if "Irregulares" in wb.sheetnames:
        wb.remove(wb["Irregulares"]) # remove qualquer possível versão antiga para evitar conflitos
    aba_irregulares_final = wb.create_sheet("Irregulares")


    colunas_irregulares_padrao = [
        "Regional", "Município", "UVR", "Técnico de UVR", 
        "Data de Envio", "Mês de referência", "Validado pelo Regional", "Observações", "Formulários para Deletar (ID)", "Validado Equipe de TI", "Resposta Equipe de TI"
    ]

    
    # Escreve o novo cabeçalho 
    for col_num, col_name in enumerate(colunas_irregulares_padrao, start=1):
        cell = aba_irregulares_final.cell(row=1, column=col_num, value=col_name)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # primeira etapa: migrar dados da aba de irregulares do arquivo de entrada
    if "Irregulares" in wb_aux.sheetnames:
        aba_irregulares_origem = wb_aux["Irregulares"]

        # cria um conjunto com as chaves de todos os novos envios para verificação
        chaves_novos_envios = set()
        for chave_composta, info in dados_atualizados.items():
            _municipio_uvr, mes_ano = chave_composta
            for data_envio in info["datas_envio"]:
                chave = (
                    normalizar_texto(info["municipio_original"]),
                    normalizar_uvr(info["uvr_nro"]),
                    data_envio,
                    mes_ano
                )
                chaves_novos_envios.add(chave)
        
        headers_origem = [cell.value for cell in aba_irregulares_origem[1]] # captura os nomes dos cabeçalhos da primeira linha da aba de origem
        try:
            # mapeia o índice de cada coluna esperada, conforme a lista de colunas padrão
            idx_map = {h: headers_origem.index(h) for h in colunas_irregulares_padrao if h in headers_origem} 
        except ValueError as e:
            print(f"AVISO: A aba 'Irregulares' em '{caminho_aux}' não tem a coluna esperada")
            idx_map = {}

        if idx_map:
            for row_origem in aba_irregulares_origem.iter_rows(min_row=2, values_only=True):
                municipio = row_origem[idx_map.get("Município")]
                if not municipio: continue

                # Cria uma chave para a linha atual do arquivo de entrada para comparação
                chave_origem = (
                    normalizar_texto(municipio), 
                    normalizar_uvr(row_origem[idx_map.get("UVR")]), 
                    row_origem[idx_map.get("Data de Envio")], 
                    row_origem[idx_map.get("Mês de referência")]
                )
                
                # migra a linha somente se a chave de origem existir nos novos envios
                if chave_origem in chaves_novos_envios:
                    idx_validado_regional = idx_map.get("Validado pelo Regional")
                    valor_validado = row_origem[idx_validado_regional] if idx_validado_regional is not None else "Não"
                    validado = "Sim" if valor_validado == "Sim" else "Não"

                    idx_validado_ti = idx_map.get("Validado Equipe de TI")
                    valor_validado_ti = row_origem[idx_validado_ti] if idx_validado_ti is not None else "Não"
                    validado_TI = "Sim" if valor_validado_ti == "Sim" else "Não"
                    
                    linha_migrada = [
                        row_origem[idx_map.get("Regional", "")] if "Regional" in idx_map else "",
                        municipio,
                        row_origem[idx_map.get("UVR", "")] if "UVR" in idx_map else "",
                        row_origem[idx_map.get("Técnico de UVR", "")] if "Técnico de UVR" in idx_map else "",
                        row_origem[idx_map.get("Data de Envio", "")] if "Data de Envio" in idx_map else "",
                        row_origem[idx_map.get("Mês de referência", "")] if "Mês de referência" in idx_map else "",
                        validado,
                        row_origem[idx_map.get("Observações", "")] if "Observações" in idx_map else "",
                        row_origem[idx_map.get("Formulários para Deletar (ID)", "")] if "Formulários para Deletar (ID)" in idx_map else "",
                        validado_TI, 
                        row_origem[idx_map.get("Resposta Equipe de TI", "")] if "Resposta Equipe de TI" in idx_map else ""                   
                    ]
                    aba_irregulares_final.append(linha_migrada)

                    # Adiciona a chave da linha migrada para evitar duplicatas na segunda etapa
                    chaves_existentes.add(chave_origem)

    # segunda etapa: adicionar novos registros irregulares do csv que ainda não existem
    for chave_composta, info in dados_atualizados.items():
        municipio_uvr, mes_ano = chave_composta
        
        if mes_ano not in wb.sheetnames and div_por_municipio.get(municipio_uvr) == nome: # verifica se é irregular
            for data_envio in info["datas_envio"]:
                chave_nova = (
                    normalizar_texto(info["municipio_original"]),
                    normalizar_uvr(info["uvr_nro"]),
                    data_envio,
                    mes_ano
                )
                
                if chave_nova not in chaves_existentes: #verifica se a chave já não existe
                    nova_linha_dados = [
                        regionais_por_municipio.get(municipio_uvr, ""),
                        info["municipio_original"], 
                        info["uvr_nro"], 
                        info["tc_uvr"],
                        data_envio, 
                        mes_ano, 
                        "Não", 
                        "", 
                        "",
                        "Não",
                        "",
                    ]
                    aba_irregulares_final.append(nova_linha_dados)
                    chaves_existentes.add(chave_nova)

    # aplicar estilização na aba de irregulares
    for row_idx in range(2, aba_irregulares_final.max_row + 1):
        for col_idx in range(1, len(colunas_irregulares_padrao) + 1):
            cell = aba_irregulares_final.cell(row=row_idx, column=col_idx)
            cell.border = bordas
            cell.alignment = alinhamento
            cell.font = Font(name='Arial', size=11)
        
        regional_cell = aba_irregulares_final.cell(row=row_idx, column=1)
        if regional_cell.value in cores_regionais:
            cor_hex = cores_regionais[regional_cell.value]
            regional_cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        status_cell = aba_irregulares_final.cell(row=row_idx, column=5)
        aplicar_estilo_status(status_cell, status_cell.value)
    
    
    if aba_irregulares_final.max_row > 1:
        # Cria o dropdown
        dv_sim_nao_irr = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)
        aba_irregulares_final.add_data_validation(dv_sim_nao_irr)
        
        # Define o range da coluna a ser afetada (H2 até a última linha)
        range_validado = f"G2:G{aba_irregulares_final.max_row}"
        range_validado_TI = f"J2:J{aba_irregulares_final.max_row}"        
        dv_sim_nao_irr.add(range_validado)
        dv_sim_nao_irr.add(range_validado_TI)

        # Define as regras de formatação condicional
        rule_sim_irr = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
        rule_nao_irr = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)


        # Aplica as regras ao range
        aba_irregulares_final.conditional_formatting.add(range_validado, rule_sim_irr)
        aba_irregulares_final.conditional_formatting.add(range_validado, rule_nao_irr)

        aba_irregulares_final.conditional_formatting.add(range_validado_TI, rule_sim_irr)
        aba_irregulares_final.conditional_formatting.add(range_validado_TI, rule_nao_irr)

    # Ajusta a largura das colunas
    if aba_irregulares_final.max_row > 1:
        for col in aba_irregulares_final.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            aba_irregulares_final.column_dimensions[col[0].column_letter].width = max_length + 5

    aba_irregulares_final.freeze_panes = 'D1'
    aba_irregulares_final.auto_filter.ref = f"A1:G1"

# --- Lógica para a Aba "Discrepantes" ---
for nome, wb in wb_final.items():
    print(f"Processando Discrepantes para '{nome}'...")
    if "Discrepantes" in wb.sheetnames:
        wb.remove(wb["Discrepantes"])
    ws_discrepantes = wb.create_sheet("Discrepantes")

    # --- 1. Construção do cabeçalho #  ---
    
    # Estilos de preenchimento
    banded_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

   
    headers = ["Regional", "Município", "UVR", "Técnico UVR", "Mês Referência", "Data de Envio", "Receita Vendas", "Média Utilizada (R$)"]
    for col_num, header_text in enumerate(headers, start=1):
        cell = ws_discrepantes.cell(row=1, column=col_num, value=header_text)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # --- 2. Coleta de dados com a lógica de semestre ---
    abas_mensais_existentes = set(wb.sheetnames)
    discrepantes_data = []

    
    # Define o ano e semestre atuais uma vez, antes do loop
    hoje = datetime.today()
    ano_atual = hoje.year
    semestre_atual = 1 if 1 <= hoje.month <= 6 else 2
   
    
    for chave_composta, info in dados_atualizados.items():
        municipio_uvr, mes_ano_envio = chave_composta
        if div_por_municipio.get(municipio_uvr) == nome and mes_ano_envio in abas_mensais_existentes:
            chave_media = (normalizar_texto(info["municipio_original"]), normalizar_uvr(info["uvr_nro"]))
            data_ref = info.get("data_referencia_dt")
            valor_envio = info.get("valor_envio")

            
            # Verifica se a data de referência é válida antes de prosseguir
            if pd.notna(data_ref):
                ano_ref = data_ref.year
                semestre_ref = 1 if 1 <= data_ref.month <= 6 else 2

                # Lógica para verificar se a data de referência está no semestre atual ou no anterior
                is_valid_semester = False
                if semestre_atual == 1:
                    # Se estamos no 1º semestre, aceita envios do 1º semestre do ano atual
                    # ou do 2º semestre do ano anterior.
                    if (ano_ref == ano_atual and semestre_ref == 1) or \
                       (ano_ref == ano_atual - 1 and semestre_ref == 2):
                        is_valid_semester = True
                else: # semestre_atual == 2
                    # Se estamos no 2º semestre, aceita envios de ambos os semestres do ano atual.
                    if ano_ref == ano_atual:
                        is_valid_semester = True
                
                # A verificação da discrepância só ocorre se o semestre for válido
                if is_valid_semester and pd.notna(valor_envio) and chave_media in dados_medias:
                    
                    if semestre_atual == 2:
                        # Se rodamos no 2º semestre, a lógica original funciona.
                        media_ref = dados_medias[chave_media]["media_s1"] if semestre_ref == 1 else dados_medias[chave_media]["media_s2"]
                    else:  # semestre_atual == 1
                        # Se rodamos no 1º semestre, a lógica precisa ser INVERTIDA.
                        media_ref = dados_medias[chave_media]["media_s2"] if semestre_ref == 1 else dados_medias[chave_media]["media_s1"]
                   
                    if pd.notna(media_ref) and media_ref > 0:
                        desvio = abs((valor_envio - media_ref) / media_ref) * 100
                        if desvio >= 40:
                            discrepantes_data.append({
                                "regional": regionais_por_municipio.get(municipio_uvr, ""),
                                "municipio": info["municipio_original"],
                                "uvr": info["uvr_nro"],
                                "tc_uvr": info.get("tc_uvr", ""),
                                "mes_ano": mes_ano_envio,
                                "data_envio": info.get("datas_envio", [""])[0],
                                "valor_envio": valor_envio,
                                "desvio": desvio,
                                "media": media_ref
                            })
       

    # --- 3. Escrita dos dados com início na linha 2 ---
    discrepantes_data.sort(key=lambda x: x["desvio"], reverse=True)

    for i, data in enumerate(discrepantes_data):
        row_idx = i + 2 # Dados agora começam na linha 2
        
        linha = [
            data["regional"], data["municipio"], data["uvr"], data["tc_uvr"],
            data["mes_ano"], data["data_envio"],
            data["valor_envio"],
            data["media"]
        ]

        ws_discrepantes.cell(row=row_idx, column=8).number_format = 'R$ #,##0.00'

        for col_idx, value in enumerate(linha, start=1):
            cell = ws_discrepantes.cell(row=row_idx, column=col_idx, value=value)
            cell.border = bordas
            cell.alignment = alinhamento
        
        # Formata a coluna de receita
        receita_cell = ws_discrepantes.cell(row=row_idx, column=7)
        receita_cell.number_format = '#,##0.00'

        # Aplica a cor do desvio APENAS na célula de Receita Vendas
        desvio_fill = None
        if data["desvio"] >= 80: desvio_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif data["desvio"] >= 60: desvio_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif data["desvio"] >= 40: desvio_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        if desvio_fill:
            receita_cell.fill = desvio_fill

    # Ajusta a largura das colunas
    for col_num in range(1, len(headers) + 1):
        col_letter = ws_discrepantes.cell(row=1, column=col_num).column_letter
        max_length = 0
        # Percorre a partir da linha 1 para incluir o cabeçalho no cálculo da largura
        for row_num in range(1, ws_discrepantes.max_row + 1):
             cell_value = ws_discrepantes.cell(row=row_num, column=col_num).value
             if cell_value:
                 max_length = max(max_length, len(str(cell_value)))
        adjusted_width = max_length + 5
        ws_discrepantes.column_dimensions[col_letter].width = adjusted_width
        
    # Congela os painéis e define o filtro (ajustado para a linha 1)
    ws_discrepantes.freeze_panes = 'A2'
    if ws_discrepantes.max_row > 1:
        ws_discrepantes.auto_filter.ref = f"A1:G{ws_discrepantes.max_row}"

    

# Salva os novos arquivos com nome atualizado
for nome, wb in wb_final.items():
    novo_caminho = pasta_scripts.parent / "form4" / f"{nome}_atualizado_form4v2.xlsx"
    wb.save(novo_caminho)
    print(f"{novo_caminho} gerado com sucesso")
