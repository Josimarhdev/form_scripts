from openpyxl import Workbook
from pathlib import Path 
import os
import sys
from dotenv import load_dotenv 

# Carrega variáveis de ambiente do arquivo .env (se existir)
load_dotenv()

# Inicializa os workbooks globais
belem_wb = Workbook()
expansao_wb = Workbook()
grs_wb = Workbook()
expansao_ms_wb = Workbook()

# Remove aba padrão de cada workbook
for wb in [belem_wb, expansao_wb, grs_wb, expansao_ms_wb]:
    wb.remove(wb.active)

# Deixa as variáveis globais disponíveis para os scripts
globals().update({
    "belem_wb": belem_wb,
    "expansao_wb": expansao_wb,
    "grs_wb": grs_wb,
    "expansao_ms_wb": expansao_ms_wb
})

print("--- Executando scripts de Forms (1 a 4) ---")
# Estes scripts são essenciais e rodam sempre
exec(open("scripts/script_form1.py").read())
exec(open("scripts/script_form2.py").read())
exec(open("scripts/script_form3.py").read())
exec(open("scripts/script_form4.py").read())


print("\n--- Verificando Integração com Banco de Dados ---")

# Lista de variáveis obrigatórias para o banco
db_vars = ["DB_NAME", "DB_USER", "DB_PASSWORD", "DB_HOST"]
# Verifica se todas as variáveis estão presentes e não são vazias
tem_credenciais = all(os.getenv(var) for var in db_vars)

if tem_credenciais:
    print("Credenciais de banco encontradas. Iniciando script de integração...")
    try:
        # Tenta executar o script de integração
        exec(open("scripts/script_integracao.py").read())
        print(">>> Integração finalizada com sucesso.")
    except Exception as e:
        # Se der erro, captura o erro e continua o fluxo
        print(f"[ERRO] Falha ao executar a integração com o banco: {e}")
        print("O script continuará para salvar os Forms restantes.")
else:
    # Se não tiver credenciais no .env, avisa e pula
    print("[AVISO] Variáveis de ambiente do banco não encontradas (DB_HOST, etc).")
    print("Pulando etapa de integração. Apenas os Forms serão gerados.")


pasta_scripts = Path(__file__).parent
pasta_saida = pasta_scripts.parent / "outputs"


mapa_saida = {
    belem_wb: "Belém",
    expansao_wb: "Expansão",
    grs_wb: "GRS",
    expansao_ms_wb: "Expansão MS"
}

print("\n--- Iniciando salvamento dos Forms 1, 2 e 3 (Monitoramento) ---")

for wb, nome_pasta in mapa_saida.items():
    # Cria o caminho completo: outputs/NomeDaPasta
    caminho_final_pasta = pasta_saida / nome_pasta
    caminho_final_pasta.mkdir(parents=True, exist_ok=True)
    
    # Define o nome do arquivo
    caminho_arquivo = caminho_final_pasta / "0 - Monitoramento Form 1, 2 e 3.xlsx"
    
    # Salva
    wb.save(caminho_arquivo)
    print(f"Salvo: {caminho_arquivo}")

print("Processo finalizado.")