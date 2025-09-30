from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import unicodedata

# Estilos reutilizáveis
cabeçalho_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
cabeçalho_font = Font(color="FFFFFF", bold=True, name='Arial', size=11)

enviado_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
enviado_font = Font(color="FFFFFF", name='Arial', size=11)
semtecnico_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
atrasado_fill = PatternFill(start_color="FF6400", end_color="FF6400", fill_type="solid")
duplicado_fill = PatternFill(start_color="A020F0", end_color="A020F0", fill_type="solid")
outras_fill = PatternFill(start_color="A020F0", end_color="A020F0", fill_type="solid")
atrasado2_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
analise_fill = PatternFill(start_color="FFA500", end_color= "FFA500", fill_type="solid")

validado_nao_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
validado_sim_fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")

cores_regionais = {
    "Gabriel": "A9C5E6",
    "Bianca": "FFFF99",
    "Valquiria": "B2FFFF",
    "Valquíria": "B2FFFF",
    "Luana": "FFCCFF",
    "Larissa": "F1E0C6",
    "Paranavaí": "9B59B6",
    "Ana Paula": "993399",
    "Londrina": "A9C5E6",
    "Francisco Beltrão": "B2FFFF",
    "Maringá": "FFCCFF",
    "Curitiba": "FFFF99",
    "Guarapuava": "F1E0C6",
    "Laranjeiras do Sul" : "FFD3AC",
    "Cibax": "FDEE00",      # Amarelo
    "Comafen": "FFA500",    # Laranja
    "Coripa": "008000",     # Verde
    "Cifra": "00BFFF",      # Azul
    "Conisul": "FF0000"     # Vermelho
}

bordas = Border(
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000")
)

alinhamento = Alignment(horizontal="center", vertical="center")

def normalizar_texto(texto):
    if isinstance(texto, str):
        texto = texto.strip().lower()
        texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return texto

def normalizar_uvr(uvr_nro):
    try:
        return str(int(uvr_nro))
    except ValueError:
        return uvr_nro


def aplicar_estilo_status(cell, status):

    if status == "Enviado": 
        cell.fill = enviado_fill
        cell.font = enviado_font
    elif status == "UVR Sem Técnico" or status == "Sem Técnico":
        cell.fill = semtecnico_fill
        cell.font = enviado_font
    elif status == "Atrasado":
        cell.fill = atrasado_fill
        cell.font = enviado_font
    elif status == "Duplicado":
        cell.fill = duplicado_fill
        cell.font = enviado_font
    elif status == "Outras Ocorrências":
        cell.fill = outras_fill
        cell.font = enviado_font
    elif status == "Atrasado >= 2":
        cell.fill = atrasado2_fill
        cell.font = enviado_font





