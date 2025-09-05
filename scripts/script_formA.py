import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from openpyxl.styles import PatternFill


try:
    from utils import (
        cabeçalho_fill, cabeçalho_font, enviado_fill, enviado_font,
        atrasado_fill, cores_regionais, bordas, alinhamento
    )
except ImportError:
    print("ERRO: O arquivo 'utils.py' não foi encontrado na pasta 'scripts'.")
    print("Por favor, certifique-se de que o seu arquivo utils.py está no lugar certo.")
    exit()



caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"
pasta_outputs = pasta_scripts.parent / "outputs"
pasta_outputs.mkdir(exist_ok=True)

caminhos_csv = {
    "belem": pasta_inputs / "formA-belém.csv",
    "expansao": pasta_inputs / "formA-expansão.csv",
    "grs": pasta_inputs / "formA-grs.csv"
}



def gerar_planilha_estilizada(df, nome_convenio):
   
    print(f"  - Criando Workbook para '{nome_convenio}'...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoramento"
    
    # Ordem das colunas e nome alterados
    colunas_finais = ['Município', 'UVR', 'Quem preencheu', 'Situação', 'Data de Envio']

    dv_status = DataValidation(type="list", formula1='"Enviado,Não Enviado"', allow_blank=True)
    ws.add_data_validation(dv_status)

    for col_num, header in enumerate(colunas_finais, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    for index, row_data in df.iterrows():
        row_num = index + 2
        for col_num, col_name in enumerate(colunas_finais, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data[col_name])
            cell.border = bordas
            cell.alignment = alinhamento

            if col_name == 'Situação':
                dv_status.add(cell)

            # Bloco de código que coloria a célula da regional foi REMOVIDO
    
    # Loop para colorir "Situação" foi ajustado para a nova posição da coluna (coluna 4)
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=4) # Coluna D é a 4ª coluna
        status = status_cell.value
        
        if status == "Enviado":
            status_cell.fill = enviado_fill
            status_cell.font = enviado_font
        elif status == "Não Enviado":
            status_cell.fill = atrasado_fill
            status_cell.font = enviado_font

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = max_length + 5

    ws.column_dimensions['A'].width = 35 # A coluna A agora é 'Município'
    ws.freeze_panes = 'A2'

    caminho_saida = pasta_outputs / f"{nome_convenio}_formA.xlsx"
    try:
        wb.save(caminho_saida)
        print(f"  - Planilha '{caminho_saida.name}' salva com sucesso!")
    except Exception as e:
        print(f"  - ERRO ao salvar '{caminho_saida.name}': {e}")



# LÓGICA PRINCIPAL DO SCRIPT

print("Iniciando o processo de geração de relatórios...")
for convenio, caminho in caminhos_csv.items():
    print(f"\nProcessando convênio: '{convenio.upper()}'")
    
    if not caminho.exists():
        print(f"  - AVISO: Arquivo '{caminho.name}' não encontrado. Pulando este convênio.")
        continue

    df = pd.read_csv(caminho, dtype=str).fillna('---')
    # Renomeando a coluna 'regional' para 'Quem preencheu'
    df.rename(columns={'regional': 'Quem preencheu', 'municipio': 'Município', 'uvr': 'UVR', 'data_envio': 'Data de Envio'}, inplace=True)
    
    # Linha que encurtava o nome da regional foi REMOVIDA
    # df['Regional'] = df['Regional'].apply(lambda x: x.split(' ')[0])

    df['Situação'] = df['Data de Envio'].apply(lambda x: 'Não Enviado' if str(x).strip() in ['---', ''] else 'Enviado')

    def formatar_data(data):
        if str(data).strip() in ['---', '']: return '---'
        try:
            return pd.to_datetime(data, format='%d/%m/%Y').strftime('%d/%m/%Y')
        except:
            return data
    df['Data de Envio'] = df['Data de Envio'].apply(formatar_data)
    
    # Definindo a ordem final das colunas no DataFrame
    df_final = df[['Município', 'UVR', 'Quem preencheu', 'Situação', 'Data de Envio']]
    
    gerar_planilha_estilizada(df_final, convenio)

print("\nProcesso finalizado.")