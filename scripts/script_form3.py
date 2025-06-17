from openpyxl import load_workbook, Workbook  # Para carregar e criar arquivos Excel
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # Para aplicar estilos nas células
from datetime import datetime  # Para lidar com datas
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from pathlib import Path  # Para manipulação de caminhos de arquivos
from utils import (  # Importa funções utilitárias e estilos personalizados
    cabeçalho_fill, cabeçalho_font, enviado_fill, enviado_font,
    semtecnico_fill, atrasado_fill, validado_nao_fill, validado_sim_fill, atrasado2_fill, outras_fill, duplicado_fill,
    cores_regionais, bordas, alinhamento,
    normalizar_texto, normalizar_uvr, aplicar_estilo_status
)

# Caminho da pasta onde o script está localizado
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"

# Caminho do arquivo do banco e arquivos auxiliares (originais do drive)
csv_file_input = pasta_inputs/"form3.csv"  
planilhas_auxiliares = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 1, 2 e 3.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 1, 2 e 3.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 1, 2 e 3.xlsx"
}


# Carrega a planilha principal
df_input = pd.read_csv(csv_file_input, dtype=str)

dados_atualizados = {}  # Dicionário para armazenar informações atualizadas

for _, row in df_input.iterrows():
    municipio = row['municipio']
    uvr_nro = row['uvr_numero']
    data_envio = row['data_envio']
    cnpj = row['cnpj']  

    # Gera a chave normalizada (municipio + UVR)
    if isinstance(municipio, str):
        municipio_uvr_normalizado = f"{normalizar_texto(municipio)}_{normalizar_uvr(uvr_nro)}"
    else:
        continue  

    # Formata a data de envio
    if isinstance(data_envio, datetime):
        data_envio_formatada = data_envio.strftime("%d/%m/%Y")
    else:
        try:
            data_envio_formatada = datetime.strptime(data_envio, "%Y-%m-%d %H:%M:%S.%f").strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_envio_formatada = ""

    # Cria a entrada no dicionário se ainda não existir
    if municipio_uvr_normalizado not in dados_atualizados:
        dados_atualizados[municipio_uvr_normalizado] = {"cnpjs": {}, "status": "Enviado"}

    # Adiciona ou atualiza as datas de envio por CNPJ
    if cnpj not in dados_atualizados[municipio_uvr_normalizado]["cnpjs"]:
        dados_atualizados[municipio_uvr_normalizado]["cnpjs"][cnpj] = [data_envio_formatada]
    else:
        dados_atualizados[municipio_uvr_normalizado]["cnpjs"][cnpj].append(data_envio_formatada)
        dados_atualizados[municipio_uvr_normalizado]["status"] = "Duplicado"  # Marca como duplicado

# Processa cada uma das planilhas auxiliares (Belém, GRS e Expansão)
for nome, caminho in planilhas_auxiliares.items():
    wb_aux = load_workbook(caminho)

    # Verifica se a aba necessária existe
    if "Form 3 - Empreendimento" not in wb_aux.sheetnames:
        print(f"A aba 'Form 3 - Empreendimento' não foi encontrada em {nome}. Nenhuma modificação será feita.")
        continue

    ws_aux = wb_aux["Form 3 - Empreendimento"]


    # Usa o workbook que foi criado no EXECUTAR_TODOS
    wb_destino = {"belem": belem_wb, "expansao": expansao_wb, "grs": grs_wb}[nome] # type: ignore
    novo_ws = wb_destino.create_sheet("Form 3 - Empreendimento")  

    dv_sim_nao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True) #dropdown com sim e nao
    novo_ws.add_data_validation(dv_sim_nao) 

    dv_status = DataValidation(type="list", formula1='"Enviado, Atrasado, Outras Ocorrências, Sem Técnico, Duplicado"', allow_blank=True) #dropdown de status
    novo_ws.add_data_validation(dv_status)     

    # Copia o cabeçalho da planilha original e aplica estilos
    headers = [cell.value for cell in ws_aux[1]]
    for col_num, header in enumerate(headers, start=1):
        cell = novo_ws.cell(row=1, column=col_num, value=header)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # Processa cada linha da planilha auxiliar
    for row_idx, row in enumerate(ws_aux.iter_rows(min_row=2, values_only=True), start=2):
        municipio_original = row[1]
        uvr_nro_original = row[2]  
        row_data = list(row)

        # Normaliza a chave para buscar nos dados atualizados
        if isinstance(municipio_original, str):
            municipio_uvr_normalizado = f"{normalizar_texto(municipio_original)}_{normalizar_uvr(uvr_nro_original)}"
            print(municipio_uvr_normalizado, ' aux')
        else:
            municipio_uvr_normalizado = ""

        # Atualiza as colunas de status e datas, se dados foram encontrados
        if municipio_uvr_normalizado in dados_atualizados:
            novas_datas = ", ".join(
                sum(dados_atualizados[municipio_uvr_normalizado]["cnpjs"].values(), [])
            )
            novo_status = dados_atualizados[municipio_uvr_normalizado]["status"]
            row_data[5] = novas_datas  # Data de envio
            row_data[4] = novo_status  # Status
        else: 
            if row_data[4] == "Sem Técnico":
                pass
            elif row_data[4] == None:
                row_data[4] = "Atrasado"

        # Colore a célula da validação
        if row_data[6] == "Não":
            novo_ws.cell(row=row_idx, column=7).fill = validado_nao_fill
        elif row_data[6] == "Sim":
            novo_ws.cell(row=row_idx, column=7).fill = validado_sim_fill

        # Aplica cor regional
        regional = row_data[0]
        if regional in cores_regionais:
            cor_hex = cores_regionais[regional]
            novo_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        # Escreve os dados na nova planilha com estilos
        for col_idx, value in enumerate(row_data, start=1):
            cell = novo_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = bordas
            cell.alignment = alinhamento
            cell.font = Font(name='Arial', size=11)

            if col_idx == 7: 
                dv_sim_nao.add(cell.coordinate) # Adiciona esta célula à regra de validação dv_sim_nao pro validado pelos regionais

            if col_idx == 10: 
                dv_sim_nao.add(cell.coordinate) # Adiciona esta célula à regra de validação dv_sim_nao pro validado pela equipe de TI  

            if col_idx == 5:
                dv_status.add(cell.coordinate) 

    # Aplica cores de preenchimento conforme o status
    for row_idx in range(2, novo_ws.max_row + 1):
        status_cell = novo_ws.cell(row=row_idx, column=5)
        status = status_cell.value
        aplicar_estilo_status(status_cell, status_cell.value)

    # Ajusta automaticamente a largura das colunas
    for col in novo_ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        novo_ws.column_dimensions[col_letter].width = max_length + 5


    if novo_ws.max_row >= 2: 

        coluna_validado_regional = f"G2:G{novo_ws.max_row}" # Coluna G é a 7ª coluna
        coluna_validado_ti = f"J2:J{novo_ws.max_row}" # Coluna G é a 10ª coluna
        coluna_status = f"E2:E{novo_ws.max_row}" # Coluna E é a 5ª coluna

        rule_sim = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
        novo_ws.conditional_formatting.add(coluna_validado_regional, rule_sim) #Se for selecionado Sim, pinta de verde
        novo_ws.conditional_formatting.add(coluna_validado_ti, rule_sim) #Se for selecionado Sim, pinta de verde

        rule_nao = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)
        novo_ws.conditional_formatting.add(coluna_validado_regional, rule_nao) #Se for selecionado Não, pinta de vermelho
        novo_ws.conditional_formatting.add(coluna_validado_ti, rule_nao) #Se for selecionado Sim, pinta de verde

        status_rules = {
            "Enviado": {"fill": enviado_fill, "font": enviado_font},
            "Atrasado": {"fill": atrasado_fill, "font": enviado_font},
            "Outras Ocorrências": {"fill": outras_fill, "font": enviado_font},
            "Sem Técnico": {"fill": semtecnico_fill, "font": enviado_font},
            "Duplicado": {"fill": duplicado_fill, "font": enviado_font}
        }

        for status_text, styles in status_rules.items():
            rule = CellIsRule(operator='equal',
                            formula=[f'"{status_text}"'],
                            stopIfTrue=True,
                            fill=styles["fill"],
                            font=styles["font"])
            novo_ws.conditional_formatting.add(coluna_status, rule)   

    # Salva a nova planilha atualizada
    #novo_caminho = pasta_scripts.parent / "form3" / f"{nome}_atualizado_form3.xlsx"
    #novo_wb.save(novo_caminho)

