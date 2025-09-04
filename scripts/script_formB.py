import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from openpyxl.styles import PatternFill

# Tenta importar as configurações de estilo do arquivo utils.py
try:
    from utils import (
        cabeçalho_fill, cabeçalho_font, enviado_fill, enviado_font,
        atrasado_fill, cores_regionais, bordas, alinhamento
    )
except ImportError:
    print("ERRO: O arquivo 'utils.py' não foi encontrado na pasta 'scripts'.")
    print("Por favor, certifique-se de que o seu arquivo utils.py está no lugar certo.")
    exit()

# --- CONFIGURAÇÃO DE CAMINHOS ---
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"
pasta_outputs = pasta_scripts.parent / "outputs"
pasta_outputs.mkdir(exist_ok=True)

# Dicionário com os caminhos para os arquivos CSV do Formulário B
caminhos_csv_formB = {
    "belem": pasta_inputs / "formB-belém.csv",
    "expansao": pasta_inputs / "formB-expansão.csv",
    "grs": pasta_inputs / "formB-grs.csv"
}

# --- FUNÇÃO DE GERAÇÃO DA PLANILHA ---

def gerar_planilha_estilizada_formB(df, nome_convenio):
    """
    Gera uma planilha Excel estilizada para o Formulário B.
    """
    print(f"  - Criando Workbook para '{nome_convenio}' (Form B)...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoramento Form B"

    # Define as colunas que aparecerão na planilha final e na ordem correta
    colunas_finais = [
        'Município', 'UVR',
        'Regional PS', 'Situação PS', 'Data Envio PS',
        'Regional LR', 'Situação LR', 'Data Envio LR',
        'Regional OS', 'Situação OS', 'Data Envio OS'
    ]

    # Aplica estilo ao cabeçalho
    for col_num, header in enumerate(colunas_finais, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # Preenche os dados e aplica estilos célula por célula
    for index, row_data in df.iterrows():
        row_num = index + 2
        for col_num, col_name in enumerate(colunas_finais, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data[col_name])
            cell.border = bordas
            cell.alignment = alinhamento

            # Colore as células de Regional de acordo com o dicionário 'cores_regionais'
            if col_name in ['Regional PS', 'Regional LR', 'Regional OS']:
                hex_color = cores_regionais.get(row_data[col_name])
                if hex_color:
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
            # Colore as células de Situação (Enviado/Não Enviado)
            if col_name in ['Situação PS', 'Situação LR', 'Situação OS']:
                if cell.value == "Enviado":
                    cell.fill = enviado_fill
                    cell.font = enviado_font
                elif cell.value == "Não Enviado":
                    cell.fill = atrasado_fill
                    cell.font = enviado_font

    # Ajusta a largura das colunas com base no conteúdo
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        # Pula o ajuste automático para a coluna B (Município) que terá largura fixa
        if column_letter == 'B':
            continue
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 5

    # Define larguras fixas para colunas específicas para melhor visualização
    ws.column_dimensions['A'].width = 35 # Município
    ws.column_dimensions['B'].width = 25 # UVR
    ws.freeze_panes = 'C2' # Congela as colunas Município e UVR

    # Salva o arquivo final
    caminho_saida = pasta_outputs / f"{nome_convenio}_formB.xlsx"
    try:
        wb.save(caminho_saida)
        print(f"  - Planilha '{caminho_saida.name}' salva com sucesso!")
    except Exception as e:
        print(f"  - ERRO ao salvar '{caminho_saida.name}': {e}")


# --- LÓGICA PRINCIPAL DO SCRIPT ---

def formatar_data(data):
    """Função auxiliar para padronizar o formato da data."""
    if str(data).strip() in ['---', '']:
        return '---'
    try:
       
        return pd.to_datetime(data, dayfirst=True).strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        return data

print("Iniciando o processo de geração de relatórios para o Formulário B...")

for convenio, caminho in caminhos_csv_formB.items():
    print(f"\nProcessando convênio: '{convenio.upper()}'")
    
    if not caminho.exists():
        print(f"  - AVISO: Arquivo '{caminho.name}' não encontrado. Pulando este convênio.")
        continue

    # Lê o CSV, garantindo que todos os dados sejam tratados como texto e preenchendo vazios
    df = pd.read_csv(caminho, dtype=str).fillna('---')

    # Renomeia as colunas do CSV para os nomes desejados na planilha
    mapa_colunas = {
        'municipio': 'Município', 'uvr': 'UVR',
        'regional_form_ps': 'Regional PS', 'data_envio_form_ps': 'Data Envio PS',
        'regional_form_lr': 'Regional LR', 'data_envio_form_lr': 'Data Envio LR',
        'regional_form_os': 'Regional OS', 'data_envio_form_os': 'Data Envio OS'
    }
    df.rename(columns=mapa_colunas, inplace=True)
    
    # Exibe apenas o primeiro nome dos regionais
    for col in ['Regional PS', 'Regional LR', 'Regional OS']:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: str(x).split(' ')[0])

    # Cria as colunas de "Situação" e formata as de "Data" para cada serviço
    servicos = [('PS', 'Data Envio PS'), ('LR', 'Data Envio LR'), ('OS', 'Data Envio OS')]
    
    for sigla, coluna_data in servicos:
        if coluna_data in df.columns:
            # Cria a coluna de Situação com base na presença de data
            df[f'Situação {sigla}'] = df[coluna_data].apply(
                lambda x: 'Não Enviado' if str(x).strip() in ['---', ''] else 'Enviado'
            )
            # Formata a coluna de data
            df[coluna_data] = df[coluna_data].apply(formatar_data)
        else:
            # Caso alguma coluna não exista no CSV, cria colunas vazias para evitar erros
            df[f'Situação {sigla}'] = 'Não Enviado'
            df[coluna_data] = '---'

    # Define a ordem final das colunas 
    ordem_final = [
        'Município', 'UVR',
        'Regional PS', 'Situação PS', 'Data Envio PS',
        'Regional LR', 'Situação LR', 'Data Envio LR',
        'Regional OS', 'Situação OS', 'Data Envio OS'
    ]
    
    # Garante que todas as colunas existam, preenchendo com '---' se necessário
    for col in ordem_final:
        if col not in df.columns:
            df[col] = '---'
            
    df_final = df[ordem_final]
    

    gerar_planilha_estilizada_formB(df_final, convenio)

print("\nProcesso finalizado.")