# Importações necessárias
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
import pandas as pd
from datetime import timedelta
from utils import (
    cabeçalho_fill, cabeçalho_font, enviado_fill, analise_fill, enviado_font,
    semtecnico_fill, atrasado_fill, validado_nao_fill, validado_sim_fill, duplicado_fill, outras_fill, atrasado2_fill,
    cores_regionais, bordas, alinhamento,
    normalizar_texto, normalizar_uvr, aplicar_estilo_status
)

# Define os caminhos dos arquivos envolvidos
caminho_script = Path(__file__).resolve()
pasta_scripts = caminho_script.parent
pasta_inputs = pasta_scripts.parent / "inputs"

# Caminho do arquivo do banco e arquivos auxiliares (originais do drive)
csv_file_input = pasta_inputs/"form4.csv"  
planilhas_auxiliares = {
    "belem": pasta_inputs / "0 - Belém" / "0 - Monitoramento Form 4.xlsx",
    "expansao": pasta_inputs / "0 - Expansão" / "0 - Monitoramento Form 4.xlsx",
    "grs": pasta_inputs / "0 - GRS II" / "0 - Monitoramento Form 4.xlsx"
}

# Carrega a planilha principal
df_input = pd.read_csv(csv_file_input, dtype=str)

# Dicionários para armazenar os dados extraídos
dados_atualizados = {}
div_por_municipio = {}
regionais_por_municipio = {}

# Converte a data de referência para o formato "MM.AA"
def converter_data_para_mes_ano(data_referencia):
    if isinstance(data_referencia, datetime):
        return data_referencia.strftime("%m.%y")
    else:
        try:
            return datetime.strptime(data_referencia, "%Y-%m-%d").strftime("%m.%y")
        except (ValueError, TypeError):
            return ""

# Remove caracteres inválidos do nome da aba
def limpar_nome_aba(nome):
    return nome.replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-").replace("?", "-").replace("[", "").replace("]", "")


for _, row in df_input.iterrows():
    municipio = row['gm_nome']
    uvr_nro = row['guvr_numero']
    data_envio = row['data_de_envio']
    tc_uvr = row['nome_tc_uvr']  
    data_referencia = row['data_de_referencia']


    if isinstance(municipio, str):
        municipio_uvr_normalizado = f"{normalizar_texto(municipio)}_{uvr_nro}"
    else:
        continue

    mes_ano = converter_data_para_mes_ano(data_referencia)

    # Tenta formatar a data de envio
    if isinstance(data_envio, datetime):
        data_envio_formatada = data_envio.strftime("%d/%m/%Y")
    else:
        try:
            data_envio_formatada = datetime.strptime(data_envio, "%Y-%m-%d").strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_envio_formatada = ""

    # Agrupa dados por município + UVR + mês/ano
    chave = (municipio_uvr_normalizado, mes_ano)
    if chave in dados_atualizados:
        dados_atualizados[chave]["datas_envio"].append(data_envio_formatada)
        dados_atualizados[chave]["status"] = "Duplicado"
    else:
        dados_atualizados[chave] = {
            "datas_envio": [data_envio_formatada],
            "status": "Enviado",
            "municipio_original": municipio,
            "uvr_nro": uvr_nro,
            "mes_ano": mes_ano,
            "tc_uvr" : tc_uvr
        }

# Cria um novo workbook para cada planilha auxiliar
wb_final = {nome: Workbook() for nome in planilhas_auxiliares}
for nome in wb_final:
    wb_final[nome].remove(wb_final[nome].active)

# Processa cada planilha auxiliar
for nome, caminho in planilhas_auxiliares.items():
    wb_aux = load_workbook(caminho)

    abas_para_copiar_completamente = ["Resumo", "Monitoramento", "Regionais"]

    for nome_aba in abas_para_copiar_completamente:
        if nome_aba in wb_aux.sheetnames:
            print(f"Copiando aba '{nome_aba}' para o arquivo de '{nome}'...")
            ws_origem = wb_aux[nome_aba]
            ws_destino = wb_final[nome].create_sheet(nome_aba)

            # Copia os dados e estilos célula por célula
            for row in ws_origem.iter_rows():
                for cell in row:
                    new_cell = ws_destino.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                        new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                        new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                        new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                        new_cell.number_format = cell.number_format

            # Copia as dimensões das colunas e linhas
            for col_letter, dim in ws_origem.column_dimensions.items():
                ws_destino.column_dimensions[col_letter].width = dim.width
            for row_index, dim in ws_origem.row_dimensions.items():
                ws_destino.row_dimensions[row_index].height = dim.height

            # Copia as células mescladas
            for merged_cell_range in ws_origem.merged_cells.ranges:
                ws_destino.merge_cells(str(merged_cell_range))
            
            for dv in ws_origem.data_validations.dataValidation:
                ws_destino.add_data_validation(dv)

            for range_string in ws_origem.conditional_formatting:
                rules_list = ws_origem.conditional_formatting[range_string]
                    
                for rule in rules_list:
                    ws_destino.conditional_formatting.add(range_string, rule)

    for aba in wb_aux.sheetnames:
        # Só processa abas no formato MM.AA
        if aba.count('.') == 1 and all(x.isdigit() for x in aba.split('.')):
            mes_ano_aux = aba
            ws_aux = wb_aux[aba]

            mes_ano_limpo = limpar_nome_aba(mes_ano_aux)
            if mes_ano_limpo not in wb_final[nome].sheetnames:
                wb_final[nome].create_sheet(title=mes_ano_limpo)

            ws_final = wb_final[nome][mes_ano_limpo]

            dv_sim_nao = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
            ws_final.add_data_validation(dv_sim_nao)

            dv_sim_nao_ti = DataValidation(type="list", formula1='"Sim,Não,Em Análise"', allow_blank=True)
            ws_final.add_data_validation(dv_sim_nao_ti)

            dv_status = DataValidation(type="list", formula1='"Enviado, Atrasado, Atrasado >= 2, Outras Ocorrências, Sem Técnico, Duplicado"', allow_blank=True) #dropdown com sim e nao
            ws_final.add_data_validation(dv_status)           

            # Copia cabeçalhos com formatação
            headers = [cell.value for cell in ws_aux[1]]
            for col_num, header in enumerate(headers, start=1):
                cell = ws_final.cell(row=1, column=col_num, value=header)
                cell.fill = cabeçalho_fill
                cell.font = cabeçalho_font
                cell.border = bordas
                cell.alignment = alinhamento

            ws_final.auto_filter.ref = f"A1:G1"

            # Calcula mês/ano atual
            hoje = datetime.today()
            mes_atual = hoje.month
            ano_atual = hoje.year

            # Função auxiliar para calcular a diferença de meses
            def diferenca_em_meses(ano_alvo, mes_alvo, ano_base, mes_base):
                return (ano_base - ano_alvo) * 12 + (mes_base - mes_alvo)

            # Processa linhas de dados
            for row_idx, row in enumerate(ws_aux.iter_rows(min_row=2, values_only=True), start=2):
                regional = row[0]
                municipio_original = row[1]
                uvr_nro_original = row[2]
                row_data = list(row)

                if not isinstance(municipio_original, str) or not municipio_original.strip():
                    continue

                municipio_uvr_normalizado = f"{normalizar_texto(municipio_original)}_{normalizar_uvr(uvr_nro_original)}"
                chave_busca = (municipio_uvr_normalizado, mes_ano_aux)
                div_por_municipio[municipio_uvr_normalizado] = nome
                regionais_por_municipio[municipio_uvr_normalizado] = regional

                situacao_atual = row_data[4]
                tem_envio_existente = bool(row_data[5]) and isinstance(row_data[5], str) and row_data[5].strip()

                # Atualiza dados conforme a planilha principal
                if chave_busca in dados_atualizados:
                    row_data[5] = ", ".join(dados_atualizados[chave_busca]["datas_envio"])
                    row_data[4] = dados_atualizados[chave_busca]["status"]
                elif not tem_envio_existente:
                    if situacao_atual not in ("UVR Sem Técnico", "Outras Ocorrências"):
                        try:
                            aba_mes, aba_ano = map(int, mes_ano_aux.split("."))
                            aba_ano += 2000
                        except:
                            aba_mes, aba_ano = None, None

                        if aba_ano and aba_mes:
                            diff = diferenca_em_meses(aba_ano, aba_mes, ano_atual, mes_atual)
                            if diff == 1:
                                row_data[4] = "Atrasado"
                            elif diff >= 2:
                                row_data[4] = "Atrasado >= 2"

                # Estiliza as linhas
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_final.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = bordas
                    cell.alignment = alinhamento
                    cell.font = Font(name='Arial', size=11)
                    
                    if col_idx == 7:
                        dv_sim_nao.add(cell.coordinate)

                    if col_idx == 10:
                        dv_sim_nao_ti.add(cell.coordinate)

                    if col_idx == 5:
                        dv_status.add(cell.coordinate) 
                    

                # Aplica cor para célula de validação
                if row_data[6] == "Não":
                    ws_final.cell(row=row_idx, column=7).fill = validado_nao_fill
                elif row_data[6] == "Sim":
                    ws_final.cell(row=row_idx, column=7).fill = validado_sim_fill

                # Aplica cor da regional
                if regional in cores_regionais:
                    cor_hex = cores_regionais[regional]
                    ws_final.cell(row=row_idx, column=1).fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

            # Aplica estilo com base no status
            for row_idx in range(2, ws_final.max_row + 1):
                status_cell = ws_final.cell(row=row_idx, column=5)
                status = status_cell.value
                aplicar_estilo_status(status_cell, status)

            # Ajusta largura das colunas
            for col in ws_final.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                col_letter = col[0].column_letter
                wb_final[nome][mes_ano_limpo].column_dimensions[col_letter].width = max_length + 5

            coluna_validado_regional = f"G2:G{ws_final.max_row}" # Coluna G é a 7ª coluna
            coluna_validado_ti = f"J2:J{ws_final.max_row}" # Coluna G é a 10ª coluna
            coluna_status = f"E2:E{ws_final.max_row}" # Coluna E é a 5ª coluna

            rule_sim = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
            ws_final.conditional_formatting.add(coluna_validado_regional, rule_sim) #Se for selecionado Sim, pinta de verde
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_sim) #Se for selecionado Sim, pinta de verde

            rule_nao = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)
            ws_final.conditional_formatting.add(coluna_validado_regional, rule_nao) #Se for selecionado Não, pinta de vermelho
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_nao) #Se for selecionado Sim, pinta de verde

            rule_analise = CellIsRule(operator='equal', formula=['"Em Análise"'], stopIfTrue=True, fill=analise_fill)  
            ws_final.conditional_formatting.add(coluna_validado_ti, rule_analise)    

            status_rules = {
            "Enviado": {"fill": enviado_fill, "font": enviado_font},
            "Atrasado": {"fill": atrasado_fill, "font": enviado_font},
            "Atrasado >= 2": {"fill": atrasado2_fill, "font": enviado_font},
            "Outras Ocorrências": {"fill": outras_fill, "font": enviado_font},
            "Sem Técnico": {"fill": semtecnico_fill, "font": enviado_font},
            "Duplicado": {"fill": duplicado_fill, "font": enviado_font}
        }

            for status_text, styles in status_rules.items():
                rule = CellIsRule(operator='equal',
                                formula=[f'"{status_text}"'],
                                stopIfTrue=True,
                                fill=styles["fill"],
                                font=styles["font"])
                ws_final.conditional_formatting.add(coluna_status, rule) 

            ws_final.freeze_panes = 'D1' #Congela as colunas A,B,C  


# Cria aba "irregulares" com registros que não se encaixam nas abas mensais
 # Processa a aba de irregulares para cada arquivo de saída (Belém, Expansão, GRS)
for nome, wb in wb_final.items():
    chaves_existentes = set()
    
    # Recarrega o workbook auxiliar de entrada para ter acesso à aba "Irregulares" original
    caminho_aux = planilhas_auxiliares[nome]
    wb_aux = load_workbook(caminho_aux)

    # Cria a nova aba de irregulares no arquivo final. Ela sempre será reconstruída
    # para garantir a migração correta e a padronização do formato.
    if "Irregulares" in wb.sheetnames:
        wb.remove(wb["Irregulares"]) # Remove qualquer versão antiga para evitar conflitos
    aba_irregulares_final = wb.create_sheet("Irregulares")


    colunas_irregulares_padrao = [
        "Regional", "Município", "UVR", "Técnico de UVR", 
        "Data de Envio", "Mês de referência", "Validado pelo Regional", "Observações"
    ]
    # Escreve o novo cabeçalho 
    for col_num, col_name in enumerate(colunas_irregulares_padrao, start=1):
        cell = aba_irregulares_final.cell(row=1, column=col_num, value=col_name)
        cell.fill = cabeçalho_fill
        cell.font = cabeçalho_font
        cell.border = bordas
        cell.alignment = alinhamento

    # ETAPA 1: Migrar dados da aba "Irregulares" do arquivo de ENTRADA
    if "Irregulares" in wb_aux.sheetnames:
        aba_irregulares_origem = wb_aux["Irregulares"]
        
        headers_origem = [cell.value for cell in aba_irregulares_origem[1]]
        try:
            idx_map = {h: headers_origem.index(h) for h in colunas_irregulares_padrao if h in headers_origem}
        except ValueError as e:
            print(f"AVISO: A aba 'Irregulares' em '{caminho_aux}' não tem a coluna esperada: {e}. A migração pode falhar.")
            idx_map = {}

        if idx_map:
            for row_origem in aba_irregulares_origem.iter_rows(min_row=2, values_only=True):
                municipio = row_origem[idx_map.get("Município")]
                if not municipio: continue

              
                validado = row_origem[idx_map.get("Validado pelo Regional")]
                if validado not in ("Sim", "Não"):
                    validado = "Não"
                
                linha_migrada = [
                    row_origem[idx_map.get("Regional", "")] if "Regional" in idx_map else "",
                    municipio,
                    row_origem[idx_map.get("UVR", "")] if "UVR" in idx_map else "",
                    row_origem[idx_map.get("Técnico de UVR", "")] if "Técnico de UVR" in idx_map else "",
                    row_origem[idx_map.get("Data de Envio", "")] if "Data de Envio" in idx_map else "",
                    row_origem[idx_map.get("Mês de referência", "")] if "Mês de referência" in idx_map else "",
                    validado,
                    row_origem[idx_map.get("Observações", "")] if "Observações" in idx_map else ""
                ]
                aba_irregulares_final.append(linha_migrada)

                chave = (
                    normalizar_texto(municipio), 
                    normalizar_uvr(row_origem[idx_map.get("UVR")]), 
                    row_origem[idx_map.get("Data de Envio")], 
                    row_origem[idx_map.get("Mês de referência")]
                )
                chaves_existentes.add(chave)

    # ETAPA 2: Adicionar novos registros irregulares do CSV que ainda não existem
    for chave_composta, info in dados_atualizados.items():
        municipio_uvr, mes_ano = chave_composta
        
        if mes_ano not in wb.sheetnames and div_por_municipio.get(municipio_uvr) == nome:
            for data_envio in info["datas_envio"]:
                chave_nova = (
                    normalizar_texto(info["municipio_original"]),
                    normalizar_uvr(info["uvr_nro"]),
                    data_envio,
                    mes_ano
                )
                
                if chave_nova not in chaves_existentes:
                    nova_linha_dados = [
                        regionais_por_municipio.get(municipio_uvr, ""),
                        info["municipio_original"], 
                        info["uvr_nro"], 
                        info["tc_uvr"],
                        data_envio, 
                        mes_ano, 
                        "Não", 
                        "", 
                    ]
                    aba_irregulares_final.append(nova_linha_dados)
                    chaves_existentes.add(chave_nova)

    # ETAPA 3: Aplicar estilos e formatação a TODAS as linhas da aba final
    for row_idx in range(2, aba_irregulares_final.max_row + 1):
        for col_idx in range(1, len(colunas_irregulares_padrao) + 1):
            cell = aba_irregulares_final.cell(row=row_idx, column=col_idx)
            cell.border = bordas
            cell.alignment = alinhamento
            cell.font = Font(name='Arial', size=11)
        
        regional_cell = aba_irregulares_final.cell(row=row_idx, column=1)
        if regional_cell.value in cores_regionais:
            cor_hex = cores_regionais[regional_cell.value]
            regional_cell.fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")

        status_cell = aba_irregulares_final.cell(row=row_idx, column=5)
        aplicar_estilo_status(status_cell, status_cell.value)
    
    
    if aba_irregulares_final.max_row > 1:
        # Cria a regra de validação (dropdown)
        dv_sim_nao_irr = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)
        aba_irregulares_final.add_data_validation(dv_sim_nao_irr)
        
        # Define o range da coluna a ser afetada (H2 até a última linha)
        range_validado = f"G2:G{aba_irregulares_final.max_row}"
        dv_sim_nao_irr.add(range_validado)

        # Define as regras de formatação condicional
        rule_sim_irr = CellIsRule(operator='equal', formula=['"Sim"'], stopIfTrue=True, fill=validado_sim_fill)
        rule_nao_irr = CellIsRule(operator='equal', formula=['"Não"'], stopIfTrue=True, fill=validado_nao_fill)

        # Aplica as regras ao range
        aba_irregulares_final.conditional_formatting.add(range_validado, rule_sim_irr)
        aba_irregulares_final.conditional_formatting.add(range_validado, rule_nao_irr)

    # Ajusta a largura das colunas
    if aba_irregulares_final.max_row > 1:
        for col in aba_irregulares_final.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            aba_irregulares_final.column_dimensions[col[0].column_letter].width = max_length + 5

    aba_irregulares_final.freeze_panes = 'D1'
    aba_irregulares_final.auto_filter.ref = f"A1:G1"

    




# Salva os novos arquivos com nome atualizado
for nome, wb in wb_final.items():
    novo_caminho = pasta_scripts.parent / "outputs" / f"{nome}_atualizado_form4.xlsx"
    wb.save(novo_caminho)
    print(f"{novo_caminho} gerado com sucesso")
